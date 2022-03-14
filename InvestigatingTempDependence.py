# -*- coding: utf-8 -*-
"""
Created on Mon Mar 14 16:13:58 2022

@author: Ethan
"""
import os
import subprocess
import openpyxl

def stringsToFloats(rawList):
    convertedList = []
    for candidate in rawList:
        try:
            convertedList.append(float(candidate))
        except:
            pass
    return convertedList

temps = [10.0, 100.0, 200.0, 300.0, 1000.0, 2000.0, 3000.0, 4000.0, 10000.0, 20000.0]
freqs = [1.0, 2.0, 3.0, 100.0, 200.0, 300.0, 400.0, 1000.0, 2000.0]
resultList = []

for freq in freqs:
    for temp in temps:
        rawOutput = subprocess.run(["./energy_sequential", str(freq), str(temp), "1000"], capture_output=True, text=True)
        output = stringsToFloats(rawOutput.stdout.split())
        resultList.append((temp, freq, output[0]))


if (os.path.exists("TempDependenceData.xlsx")):
    workbook = openpyxl.load_workbook("TempDependenceData.xlsx")
else:
    workbook = openpyxl.Workbook()

currentSheet = workbook.active

nextRow = currentSheet.max_row + 1

for index, result in enumerate(resultList):
    currentSheet.cell(column = 2, row = nextRow).value = result[0]
    currentSheet.cell(column = 3, row = nextRow).value = result[1]
    currentSheet.cell(column = 4, row = nextRow).value = result[2]
    nextRow += 1

workbook.save("TempDependenceData.xlsx")